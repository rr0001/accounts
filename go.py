# -*- coding: utf-8 -*-

import locale
import os
import re
import warnings
from pprint import pprint

from dotenv import load_dotenv
from openpyxl import load_workbook

from accounts import update_form_values


def f(number: float):
    """Returns number formatted with thousands separator with 2 decimal points"""
    return locale.format_string("%.2f", number, grouping=True) if number else ""


if __name__ == "__main__":

    load_dotenv()

    locale.setlocale(locale.LC_ALL, os.getenv("locale", "en"))
    in_folder = os.getenv("source_folder")
    out_folder = os.getenv("out_folder")
    pdf_file = os.getenv("pdf")
    pdf_file_name = pdf_file

    infile = f"{in_folder}/{pdf_file_name}"

    xlsx_file = os.getenv("xlsx")
    xlsx_file_name = f"{in_folder}/{xlsx_file}"

    # suppress the user warnings when opening the xlsx file
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=UserWarning)
        wb = load_workbook(xlsx_file_name, data_only=True)
        # ws = wb.active

    for ws_name in wb.sheetnames:
        outfile = f"{out_folder}/{ws_name}_{pdf_file_name}"

        ws = wb[ws_name]

        table_ref = None
        table_names = list(ws.tables.keys())
        for table in table_names:
            if ws[ws.tables[table].ref.split(":")[0]].value == "FECHA":
                table_ref = ws.tables[table].ref

        # table_ref = ws.tables["Table1"].ref
        # table = ws[table_ref]

        if table_ref is None:
            raise Exception("Unable to find the correct table with data.")

        row_start = 2
        row_end = int(table_ref.split(":")[1][1:])

        year = ws.cell(row=8, column=18).value
        month = ws.cell(row=7, column=18).value
        month_end = ws.cell(row=6, column=18).value

        # get totals
        re_anterior = ws.cell(row=2, column=18).value  # recibido saldo anterior
        cp_anterior = ws.cell(row=3, column=18).value  # cuenta principal saldo anterior
        oo_anterior = ws.cell(row=4, column=18).value  # "otra" saldo anterior

        re_ent_total = ws.cell(row=row_end, column=4).value
        re_sal_total = ws.cell(row=row_end, column=5).value
        re_sal_final = (
            float(re_anterior or 0)
            + float(re_ent_total or 0)
            - float(re_sal_total or 0)
        )
        cp_ent_total = ws.cell(row=row_end, column=6).value
        cp_sal_total = ws.cell(row=row_end, column=7).value
        cp_sal_final = (
            float(cp_anterior or 0)
            + float(cp_ent_total or 0)
            - float(cp_sal_total or 0)
        )

        oo_ent_total = ws.cell(row=row_end, column=8).value
        oo_sal_total = ws.cell(row=row_end, column=9).value
        oo_sal_final = (
            float(oo_anterior or 0)
            + float(oo_ent_total or 0)
            - float(oo_sal_total or 0)
        )

        total = re_sal_final + cp_sal_final + oo_sal_final

        # format the totals in locale
        re_ent_total_fmt = f(re_ent_total)
        re_sal_total_fmt = f(re_sal_total)
        cp_ent_total_fmt = f(cp_ent_total)
        cp_sal_total_fmt = f(cp_sal_total)
        oo_ent_total_fmt = f(oo_ent_total)
        oo_sal_total_fmt = f(oo_sal_total)

        re_anterior_fmt = f(re_anterior)
        cp_anterior_fmt = f(cp_anterior)
        oo_anterior_fmt = f(oo_anterior)

        re_sal_final_fmt = f(re_sal_final)
        cp_sal_final_fmt = f(cp_sal_final)
        oo_sal_final_fmt = f(oo_sal_final)

        total_fmt = f(total)

        # conciliation
        col = 20
        reconcile_date = ws.cell(row=28, column=col).value or month_end
        row_1 = ws.cell(row=29, column=col).value
        row_2 = ws.cell(row=30, column=col).value
        row_3 = ws.cell(row=31, column=col).value
        row_4 = ws.cell(row=32, column=col).value
        # row_5_line_1 = ws.cell(row=32, column=col).value
        # row_5_line_2 = ws.cell(row=33, column=col).value
        # row_5_line_3 = ws.cell(row=34, column=col).value
        # row_5_line_4 = ws.cell(row=35, column=col).value
        # row_5_line_5 = ws.cell(row=36, column=col).value
        # row_5_line_6 = ws.cell(row=37, column=col).value
        # row_5_line_7 = ws.cell(row=38, column=col).value
        row_6 = ws.cell(row=33, column=col).value
        row_7 = ws.cell(row=34, column=col).value
        row_8 = ws.cell(row=35, column=col).value
        row_9 = ws.cell(row=36, column=col).value

        row_1_fmt = f(row_1)
        row_2_fmt = f(row_2)
        row_3_fmt = f(row_3)
        row_4_fmt = f(row_4)
        row_6_fmt = f(row_6)
        row_7_fmt = f(row_7)
        row_8_fmt = f(row_8)
        row_9_fmt = f(row_9)

        # prepare the data_dict with initial values
        data_dict = {
            "900_1_Text_C": os.getenv("cong"),
            "900_2_Text_C": os.getenv("city"),
            "900_3_Text_C": os.getenv("state"),
            "900_4_Text_C": month,
            "900_5_Text_C": year,
            "901_53_S26TotalValue": re_ent_total_fmt,
            "901_106_S26TotalValue": re_sal_total_fmt,
            "902_53_S26TotalValue": cp_ent_total_fmt,
            "902_106_S26TotalValue": cp_sal_total_fmt,
            "903_53_S26TotalValue": oo_ent_total_fmt,
            "903_106_S26TotalValue": oo_sal_total_fmt,
            "904_28_Text_C": month_end,  # month end
            "904_29_S26Amount": re_anterior_fmt,  # re_saldo_anterior FILL IN
            "904_30_S26TotalAmount": re_ent_total_fmt,  # re_entrada
            "904_31_S26TotalAmount": re_sal_total_fmt,  # re_salida
            "904_32_S26TotalAmount": re_sal_final_fmt,  # re_salida
            "904_33_S26Amount": cp_anterior_fmt,  # cp_saldo_anterior FILL IN
            "904_34_S26TotalAmount": cp_ent_total_fmt,  # cp_entrada
            "904_35_S26TotalAmount": cp_sal_total_fmt,  # cp_salida
            "904_36_S26TotalAmount": cp_sal_final_fmt,  # cp_salida
            "904_38_S26Amount": oo_anterior_fmt,  # oo_saldo_anterior FILL IN
            "904_39_S26TotalAmount": oo_ent_total_fmt,  # oo_entrada
            "904_40_S26TotalAmount": oo_sal_total_fmt,  # oo_salida
            "904_41_S26TotalAmount": oo_sal_final_fmt,  # oo_salida
            "904_42_S26TotalAmount": total_fmt,
            "904_1_Text_C": reconcile_date,
            "904_2_S26Amount": row_1_fmt,
            "904_3_S26Amount": row_2_fmt,
            "904_4_S26Amount": row_3_fmt,
            "904_5_S26TotalAmount": row_4_fmt,
            "904_20_S26TotalAmount": row_6_fmt,
            "904_21_S26Amount": row_7_fmt,
            "904_22_S26Amount": row_8_fmt,
            "904_23_S26TotalAmount": row_9_fmt,
        }

        # columns to add to the data_dict
        cols = [
            {"name": "date", "field": "900_7_Text_C", "col_num": 1},
            {"name": "descr", "field": "900_59_Text", "col_num": 2},
            {"name": "tc", "field": "900_111_Text_C", "col_num": 3},
            {"name": "in1", "field": "901_1_S26Value", "col_num": 4},
            {"name": "out1", "field": "901_54_S26Value", "col_num": 5},
            {"name": "in2", "field": "902_1_S26Value", "col_num": 6},
            {"name": "out2", "field": "902_54_S26Value", "col_num": 7},
            {"name": "in3", "field": "903_1_S26Value", "col_num": 8},
            {"name": "out3", "field": "903_54_S26Value", "col_num": 9},
        ]

        # add the columns to the data_dict
        for col in cols:
            field = col["field"]
            for n in range(row_start, row_end):
                cell_data = ws.cell(row=n, column=col["col_num"]).value
                if cell_data:
                    if type(cell_data) in [int, float]:
                        data_dict[field] = f(cell_data)  # col + "_" + str(n)
                    elif cell_data.strip() != "":
                        data_dict[field] = cell_data
                field = re.sub(
                    r"_([^_][0-9]*)_",
                    lambda x: f"_{str(int(x.group(1))+1).zfill(len(x.group(1)))}_",
                    field,
                )
        final_dict = {k: v for k, v in data_dict.items() if v != ""}
        # pprint(final_dict)
        update_form_values(
            infile,
            outfile,
            final_dict,
        )
